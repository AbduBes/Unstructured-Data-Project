import openpyxl
import chardet
import os
from datetime import datetime

import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'src')))

from storage.mongo import save_to_mongo
from utils.logger import logging


def detect_encoding(file_path):
    """Detect file encoding using chardet."""
    with open(file_path, "rb") as f:
        raw = f.read()
    result = chardet.detect(raw)
    encoding = result.get("encoding", "utf-8") or "utf-8"
    logging.info(f"Detected encoding for {file_path}: {encoding}")
    return encoding


def extract_data_from_excel(file_path):
    """Extract data from all sheets in an Excel file, including formula cells."""
    sheets_data = []

    try:
        # data_only=True reads cached formula values instead of formula strings
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []

            for row in ws.iter_rows(values_only=True):
                # Convert all cell values to strings, handle None
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                # Skip entirely empty rows
                if any(cell.strip() for cell in cleaned_row):
                    rows.append(cleaned_row)

            sheets_data.append({
                "sheet_name": sheet_name,
                "rows": rows,
                "row_count": len(rows),
                "col_count": ws.max_column
            })

        logging.info(f"Extracted {len(sheets_data)} sheet(s) from Excel: {file_path}")

    except Exception as e:
        logging.error(f"Failed to extract Excel file {file_path}: {e}")

    return sheets_data


def process_excel(file_path):
    """Extract, structure, and store Excel content in MongoDB."""
    file_name = os.path.basename(file_path)
    logging.info(f"Processing Excel file: {file_name}")

    sheets = extract_data_from_excel(file_path)

    for sheet in sheets:
        document = {
            "content": {
                "sheet_name": sheet["sheet_name"],
                "rows": sheet["rows"],
                "row_count": sheet["row_count"],
                "col_count": sheet["col_count"]
            },
            "metadata": {
                "file_name": file_name,
                "document_type": "excel",
                "extraction_timestamp": datetime.utcnow(),
                "source": file_path,
                "extraction_library": "openpyxl",
                "sheet_name": sheet["sheet_name"]
            }
        }

        save_to_mongo(document, source_url=file_path)
        logging.info(f"Stored sheet '{sheet['sheet_name']}' from {file_name} in MongoDB")


def process_all_excel_files(folder_path="data/raw/excel"):
    """Process all .xlsx files in a folder."""
    if not os.path.exists(folder_path):
        logging.warning(f"Excel folder not found: {folder_path}")
        return

    excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    if not excel_files:
        logging.warning(f"No Excel files found in {folder_path}")
        return

    for filename in excel_files:
        file_path = os.path.join(folder_path, filename)
        process_excel(file_path)


if __name__ == "__main__":
    process_all_excel_files()